"""
Реальная генерация PDF/XLSX и отправка отчётов по SMTP.
"""
from __future__ import annotations

import io
import logging
import os
import smtplib
from datetime import datetime, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import List, Optional, Tuple

import requests

logger = logging.getLogger(__name__)

REPORTS_DIR = Path(os.getenv("REPORTS_STORAGE_DIR", "/tmp/agro_reports"))
FIELD_SERVICE_URL = os.getenv("FIELD_SERVICE_URL", "http://field-service:8082")
INTERNAL_API_TOKEN = os.getenv("INTERNAL_API_TOKEN", "")

MAIL_HOST = os.getenv("MAIL_HOST", "").strip()
MAIL_PORT = int(os.getenv("MAIL_PORT", "587") or "587")
MAIL_USERNAME = os.getenv("MAIL_USERNAME", "").strip()
MAIL_PASSWORD = os.getenv("MAIL_PASSWORD", "").strip()
MAIL_FROM = (os.getenv("MAIL_FROM", "").strip() or MAIL_USERNAME)

DEJAVU_PATHS = (
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf",
)


def _field_headers() -> Optional[dict]:
    if not INTERNAL_API_TOKEN:
        return None
    return {"X-Internal-Token": INTERNAL_API_TOKEN}


def fetch_fields_for_report() -> List[dict]:
    hdrs = _field_headers()
    for base in (FIELD_SERVICE_URL, "http://field-service:8082", "http://localhost:8082"):
        try:
            r = requests.get(f"{base.rstrip('/')}/api/fields", headers=hdrs, timeout=12)
            if r.ok:
                data = r.json()
                if isinstance(data, list):
                    return data
        except Exception as e:
            logger.debug("fetch_fields_for_report %s: %s", base, e)
    return []


def _escape_rl(text: str) -> str:
    return (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _register_reportlab_font() -> str:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    for p in DEJAVU_PATHS:
        path = Path(p)
        if path.is_file():
            try:
                pdfmetrics.registerFont(TTFont("AgroDejaVu", str(path)))
                return "AgroDejaVu"
            except Exception:
                logger.warning("Could not register font %s", p)
    return "Helvetica"


def build_pdf_bytes(report_name: str, template_id: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    fields = fetch_fields_for_report()
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    font = _register_reportlab_font()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.8 * cm,
        leftMargin=1.8 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="T", parent=styles["Heading1"], fontName=font, fontSize=16, leading=20)
    body_style = ParagraphStyle(name="B", parent=styles["Normal"], fontName=font, fontSize=10, leading=14)

    story = [
        Paragraph(_escape_rl(report_name), title_style),
        Spacer(1, 8),
        Paragraph(_escape_rl(f"Сформировано: {now}"), body_style),
        Paragraph(_escape_rl(f"Шаблон (ID): {template_id or '—'}"), body_style),
        Spacer(1, 14),
        Paragraph(_escape_rl("Сводка по полям (данные field-service):"), body_style),
        Spacer(1, 6),
    ]

    headers = ["Поле", "Площадь, га", "Культура", "Статус", "Влажность %"]
    rows: List[List[str]] = [headers]
    for f in fields:
        rows.append([
            str(f.get("name", "—")),
            str(f.get("area", "—")),
            str(f.get("cropType", "—")),
            str(f.get("status", "—")),
            str(f.get("currentMoistureLevel", "") if f.get("currentMoistureLevel") is not None else "—"),
        ])
    if len(rows) == 1:
        rows.append(["Нет данных: field-service недоступен или список пуст", "", "", "", ""])

    col_widths = [5.5 * cm, 2.2 * cm, 2.8 * cm, 2.2 * cm, 2.5 * cm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, -1), font),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f8f9fa")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 16))
    story.append(
        Paragraph(
            _escape_rl("Документ сформирован сервисом analytics-service (AgroAnalytics)."),
            body_style,
        )
    )

    doc.build(story)
    return buf.getvalue()


def build_xlsx_bytes(report_name: str, template_id: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    fields = fetch_fields_for_report()
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.append([report_name])
    ws.append([f"Сформировано: {now}"])
    ws.append([f"Шаблон (ID): {template_id or '—'}"])
    ws.append([])
    ws.append(["Поле", "Площадь, га", "Культура", "Статус", "Влажность %"])
    header_row = ws.max_row
    for f in fields:
        mo = f.get("currentMoistureLevel")
        ws.append([
            f.get("name", "—"),
            f.get("area", "—"),
            f.get("cropType", "—"),
            f.get("status", "—"),
            mo if mo is not None else "—",
        ])
    if not fields:
        ws.append(["Нет данных полей (field-service)", "", "", "", ""])

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A73E8")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_report_bytes(report_name: str, template_id: str, fmt: str) -> Tuple[bytes, str, str]:
    """Возвращает (bytes, filename, media_type)."""
    fmt = (fmt or "pdf").lower()
    if fmt == "excel":
        fmt = "xlsx"
    if fmt == "xlsx":
        data = build_xlsx_bytes(report_name, template_id)
        safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in report_name)[:80]
        return data, f"{safe or 'report'}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    data = build_pdf_bytes(report_name, template_id)
    safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in report_name)[:80]
    return data, f"{safe or 'report'}.pdf", "application/pdf"


def ensure_reports_dir() -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return REPORTS_DIR


def write_report_to_disk(report_id: str, fmt: str, content: bytes) -> Path:
    ensure_reports_dir()
    ext = "pdf" if (fmt or "").lower() == "pdf" else "xlsx"
    path = REPORTS_DIR / f"{report_id}.{ext}"
    path.write_bytes(content)
    return path


def read_report_from_disk(report_id: str, fmt: str) -> Optional[bytes]:
    ext = "pdf" if (fmt or "").lower() == "pdf" else "xlsx"
    path = REPORTS_DIR / f"{report_id}.{ext}"
    if path.is_file():
        return path.read_bytes()
    return None


def send_report_via_email(
    recipients: List[str],
    subject: str,
    body_text: str,
    attachment: bytes,
    attachment_filename: str,
) -> None:
    if not MAIL_HOST:
        raise RuntimeError("SMTP не настроен: задайте MAIL_HOST (и при необходимости MAIL_USERNAME, MAIL_PASSWORD) в окружении analytics-service")
    if not MAIL_FROM:
        raise RuntimeError("Задайте MAIL_FROM или MAIL_USERNAME для отправки почты")

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = MAIL_FROM
    msg["To"] = ", ".join(recipients)
    msg.attach(MIMEText(body_text, "plain", "utf-8"))

    part = MIMEApplication(attachment, _subtype="octet-stream")
    part.add_header("Content-Disposition", "attachment", filename=attachment_filename)
    msg.attach(part)

    with smtplib.SMTP(MAIL_HOST, MAIL_PORT, timeout=45) as smtp:
        smtp.ehlo()
        try:
            smtp.starttls()
            smtp.ehlo()
        except smtplib.SMTPException:
            pass
        if MAIL_USERNAME:
            smtp.login(MAIL_USERNAME, MAIL_PASSWORD) 
        smtp.sendmail(MAIL_FROM, recipients, msg.as_string())
